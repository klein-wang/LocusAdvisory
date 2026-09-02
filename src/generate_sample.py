import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict


def generate_sample_excel(output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wealth Data"

    from sow_types import SOW_TYPES

    months = _generate_months("2025-01", 12)

    headers = ["SOW Name", "SOW Type"] + months
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    sample_data = _build_sample_data(months)

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx <= 2:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                if isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0.00'

    col_widths = [22, 20] + [12] * len(months)
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = width

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def _generate_months(start: str, count: int) -> List[str]:
    from datetime import datetime
    result = []
    year, month = int(start[:4]), int(start[5:7])
    for _ in range(count):
        result.append(f"{year:04d}-{month:02d}")
        month += 1
        if month > 12:
            month = 1
            year += 1
    return result


def _build_sample_data(months: List[str]) -> List[list]:
    def salary(base, months_list):
        vals = []
        for i in range(len(months_list)):
            vals.append(round(base * (1 + 0.025) ** (i / 12), 2))
        return vals

    def freelance(base, months_list):
        vals = []
        import random
        for i in range(len(months_list)):
            factor = 1.0 + (i * 0.03) + random.uniform(-0.1, 0.15)
            vals.append(round(base * factor, 2))
        return vals

    def compound(initial, annual_rate, months_list):
        vals = []
        monthly_rate = (1 + annual_rate) ** (1 / 12) - 1
        val = initial
        for _ in range(len(months_list)):
            vals.append(round(val, 2))
            val *= (1 + monthly_rate)
        return vals

    def compound_with_contrib(initial, annual_rate, monthly_contrib, months_list):
        vals = []
        monthly_rate = (1 + annual_rate) ** (1 / 12) - 1
        val = initial
        for _ in range(len(months_list)):
            vals.append(round(val, 2))
            val = val * (1 + monthly_rate) + monthly_contrib
        return vals

    def linear_decrease(initial, monthly_decrease, months_list):
        vals = []
        val = initial
        for _ in range(len(months_list)):
            vals.append(round(val, 2))
            val = val - monthly_decrease
        return vals

    import random
    random.seed(42)

    rows = []

    sows = [
        ("Primary Salary", "income", salary(8500, months)),
        ("Freelance Income", "income", freelance(2000, months)),
        ("Index Fund", "investment", compound_with_contrib(25000, 0.07, 500, months)),
        ("Individual Stocks", "investment", compound(8000, 0.06, months)),
        ("401(k) Retirement", "retirement", compound_with_contrib(15000, 0.07, 1200, months)),
        ("Rental Property", "real_estate", compound(250000, 0.04, months)),
        ("Primary Residence", "real_estate", compound(750000, 0.03, months)),
        ("Bitcoin Holdings", "crypto", compound(15000, 0.10, months)),
        ("Emergency Fund", "cash", compound(12000, 0.03, months)),
        ("Credit Card Balance", "credit", linear_decrease(-3500, 300, months)),
        ("Auto Loan", "credit", linear_decrease(-18000, 800, months)),
    ]

    for name, sow_type, values in sows:
        rows.append([name, sow_type] + values)

    return rows


if __name__ == "__main__":
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_dir = os.path.join(project_root, "output")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "sample_wealth_data.xlsx")
    result = generate_sample_excel(path)
    print(f"Sample Excel generated: {result}")