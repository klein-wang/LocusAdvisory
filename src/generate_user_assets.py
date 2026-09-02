import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def compute_backward_values_with_contrib(current_value, annual_rate, monthly_contrib, months_back):
    monthly_rate = (1 + annual_rate) ** (1 / 12) - 1
    values = []
    for n in range(months_back, 0, -1):
        val = (current_value - monthly_contrib) / (1 + monthly_rate)
        current_value = val
        values.insert(0, round(val, 2))
    return values


def compute_backward_values(current_value, annual_rate, months_back):
    monthly_rate = (1 + annual_rate) ** (1 / 12) - 1
    values = []
    for n in range(months_back, 0, -1):
        val = current_value / ((1 + monthly_rate) ** n)
        values.append(round(val, 2))
    return values


def generate_user_asset_excel(output_path):
    months = ["2026-05", "2026-06", "2026-07", "2026-08"]

    net_monthly_savings = 42000 - 7000 - 10000

    hsbc_backward = compute_backward_values_with_contrib(
        239438.69, 0.035, net_monthly_savings, 3
    )
    hsbc_values = hsbc_backward + [239438.69]

    cmwb_backward = compute_backward_values(12770.20, 0.03, 3)
    cmwb_values = cmwb_backward + [12770.20]

    lb_backward = compute_backward_values(160000, 0.045, 3)
    lb_values = lb_backward + [160000]

    rows = [
        ["Cash Wallet", "cash", 30, 30, 30, 30],
        ["Octopus", "cash", 59, 59, 59, 59],
        ["HSBC Savings", "savings"] + hsbc_values,
        ["CMWB Savings", "savings"] + cmwb_values,
        ["Longbridge Investment", "investment"] + lb_values,
        ["Earnmore Credit", "credit", -310, -310, -310, -310],
        ["HS Travel Credit", "credit", -7357, -7357, -7357, -7357],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "HKD Assets"

    headers = ["SOW Name", "SOW Type"] + months

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx <= 2:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                if isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0.00'

    col_widths = [28, 20] + [14] * len(months)
    for i, width in enumerate(col_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = width

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_dir = os.path.join(project_root, "output")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, "user_hkd_assets.xlsx")
    result = generate_user_asset_excel(path)
    print(f"User asset Excel generated: {result}")