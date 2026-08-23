from typing import Dict, List
from datetime import datetime

from sow_types import SOW_TYPES, get_sow_type


class SOWData:
    def __init__(
        self,
        name: str,
        sow_type: str,
        monthly_values: Dict[str, float],
    ):
        self.name = name
        self.sow_type = sow_type
        self.monthly_values = monthly_values

    @property
    def type_label(self) -> str:
        return get_sow_type(self.sow_type).label

    @property
    def months(self) -> List[str]:
        return sorted(self.monthly_values.keys())

    @property
    def latest_month(self) -> str:
        months = self.months
        return months[-1] if months else None

    @property
    def latest_value(self) -> float:
        month = self.latest_month
        return self.monthly_values.get(month, 0.0) if month else 0.0

    def get_value(self, month: str) -> float:
        return self.monthly_values.get(month, 0.0)

    def to_dict(self) -> dict:
        return {
            "name": self.name,
            "sow_type": self.sow_type,
            "type_label": self.type_label,
            "monthly_values": self.monthly_values,
        }


def load_excel(file_path: str) -> List[SOWData]:
    try:
        return _load_via_pandas(file_path)
    except Exception as pandas_err:
        try:
            return _load_via_openpyxl(file_path)
        except FileNotFoundError:
            raise
        except Exception as openpyxl_err:
            raise type(openpyxl_err)(str(openpyxl_err)) from openpyxl_err


def _load_via_pandas(file_path: str) -> List[SOWData]:
    import pandas as pd

    df = pd.read_excel(file_path)
    return _parse_dataframe(df)


def _load_via_openpyxl(file_path: str) -> List[SOWData]:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Excel file must have at least a header row and one data row.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    required_cols = {"SOW Name", "SOW Type"}
    actual_cols = set(h for h in headers if h)
    missing = required_cols - actual_cols
    if missing:
        raise ValueError(
            f"Excel is missing required columns: {missing}. "
            f"Found columns: {headers}"
        )

    name_idx = headers.index("SOW Name")
    type_idx = headers.index("SOW Type")

    monthly_cols = [
        (i, h)
        for i, h in enumerate(headers)
        if h not in ("SOW Name", "SOW Type") and h
    ]

    sow_list: List[SOWData] = []
    for row in rows[1:]:
        if not any(row):
            continue

        name = str(row[name_idx] if name_idx < len(row) else "").strip()
        if not name:
            continue

        sow_type = str(row[type_idx] if type_idx < len(row) else "").strip().lower()

        if sow_type not in SOW_TYPES:
            raise ValueError(
                f"Row '{name}': invalid SOW type '{sow_type}'. "
                f"Available types: {', '.join(SOW_TYPES.keys())}"
            )

        monthly_values: Dict[str, float] = {}
        for col_idx, col_name in monthly_cols:
            if col_idx < len(row):
                val = row[col_idx]
                if val is not None and val == val:
                    try:
                        monthly_values[str(col_name).strip()] = float(val)
                    except (ValueError, TypeError):
                        pass

        sow_list.append(
            SOWData(
                name=name,
                sow_type=sow_type,
                monthly_values=monthly_values,
            )
        )

    return sow_list


def _parse_dataframe(df) -> List[SOWData]:
    import pandas as pd

    required_cols = {"SOW Name", "SOW Type"}
    actual_cols = set(df.columns)
    missing = required_cols - actual_cols
    if missing:
        raise ValueError(
            f"Excel is missing required columns: {missing}. "
            f"Found columns: {list(df.columns)}"
        )

    monthly_cols = [
        c for c in df.columns if c not in ("SOW Name", "SOW Type")
    ]
    monthly_cols = sorted(monthly_cols)

    sow_list: List[SOWData] = []
    for _, row in df.iterrows():
        name = str(row["SOW Name"]).strip()
        sow_type = str(row["SOW Type"]).strip().lower()

        if sow_type not in SOW_TYPES:
            raise ValueError(
                f"Row '{name}': invalid SOW type '{sow_type}'. "
                f"Available types: {', '.join(SOW_TYPES.keys())}"
            )

        monthly_values: Dict[str, float] = {}
        for col in monthly_cols:
            val = row[col]
            if pd.notna(val):
                monthly_values[str(col).strip()] = float(val)

        sow_list.append(
            SOWData(
                name=name,
                sow_type=sow_type,
                monthly_values=monthly_values,
            )
        )

    return sow_list


def get_month_range(
    start_month: str, months: int
) -> List[str]:
    try:
        start = datetime.strptime(start_month, "%Y-%m")
    except ValueError:
        start = datetime.strptime(start_month, "%Y-%m")

    result = []
    year = start.year
    month = start.month
    for _ in range(months):
        result.append(f"{year:04d}-{month:02d}")
        month += 1
        if month > 12:
            month = 1
            year += 1
    return result


def validate_month_format(month_str: str) -> bool:
    try:
        datetime.strptime(month_str, "%Y-%m")
        return True
    except ValueError:
        return False