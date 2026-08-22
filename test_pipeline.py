import os
import sys
import json
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from sow_types import SOW_TYPES, get_sow_type, list_sow_types
from excel_parser import load_excel, SOWData, get_month_range, validate_month_format
from forecast_engine import (
    forecast_sow,
    forecast_portfolio,
    get_forecast_months,
)
from breakdown import (
    compute_monthly_totals,
    compute_percentage_breakdown,
    compute_type_breakdown,
    compute_asset_vs_liability,
    compute_net_worth_growth,
)
from main import run_pipeline


class TestSOWTypes(unittest.TestCase):
    def test_sow_types_exist(self):
        self.assertIn("cash", SOW_TYPES)
        self.assertIn("income", SOW_TYPES)
        self.assertIn("investment", SOW_TYPES)
        self.assertIn("retirement", SOW_TYPES)
        self.assertIn("real_estate", SOW_TYPES)
        self.assertIn("crypto", SOW_TYPES)
        self.assertIn("personal_property", SOW_TYPES)
        self.assertIn("credit", SOW_TYPES)

    def test_get_sow_type(self):
        t = get_sow_type("investment")
        self.assertEqual(t.key, "investment")
        self.assertTrue(t.is_asset)
        self.assertEqual(t.default_annual_growth, 0.07)

    def test_get_sow_type_invalid(self):
        with self.assertRaises(ValueError):
            get_sow_type("nonexistent")

    def test_list_sow_types(self):
        types = list_sow_types()
        self.assertIsInstance(types, list)
        self.assertGreater(len(types), 0)


class TestExcelParser(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.tmpdir, "test_data.xlsx")
        self._create_test_excel()

    def _create_test_excel(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Test"

        months = ["2025-01", "2025-02", "2025-03"]
        headers = ["SOW Name", "SOW Type"] + months
        ws.append(headers)

        ws.append(["Salary", "income", 5000, 5100, 5200])
        ws.append(["Index Fund", "investment", 20000, 21000, 22000])
        ws.append(["Credit Card", "credit", -3000, -2800, -2600])

        wb.save(self.excel_path)

    def test_load_excel(self):
        sow_list = load_excel(self.excel_path)
        self.assertEqual(len(sow_list), 3)
        self.assertEqual(sow_list[0].name, "Salary")
        self.assertEqual(sow_list[0].sow_type, "income")
        self.assertEqual(sow_list[1].name, "Index Fund")
        self.assertEqual(sow_list[2].name, "Credit Card")

    def test_sow_data_properties(self):
        sow_list = load_excel(self.excel_path)
        salary = sow_list[0]
        self.assertEqual(salary.latest_month, "2025-03")
        self.assertEqual(salary.latest_value, 5200)
        self.assertEqual(salary.get_value("2025-01"), 5000)
        self.assertEqual(salary.type_label, "Stable Income")

    def test_load_excel_invalid_path(self):
        with self.assertRaises(FileNotFoundError):
            load_excel("/nonexistent/path.xlsx")

    def test_load_excel_missing_columns(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Wrong Col", "2025-01"])
        ws.append(["data", 100])
        path = os.path.join(self.tmpdir, "missing_cols.xlsx")
        wb.save(path)
        with self.assertRaises(ValueError):
            load_excel(path)

    def test_load_excel_invalid_sow_type(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["SOW Name", "SOW Type", "2025-01"])
        ws.append(["Test", "invalid_type", 100])
        path = os.path.join(self.tmpdir, "invalid_type.xlsx")
        wb.save(path)
        with self.assertRaises(ValueError):
            load_excel(path)

    def test_get_month_range(self):
        months = get_month_range("2025-11", 5)
        self.assertEqual(len(months), 5)
        self.assertEqual(months[0], "2025-11")
        self.assertEqual(months[1], "2025-12")
        self.assertEqual(months[2], "2026-01")
        self.assertEqual(months[4], "2026-03")

    def test_validate_month_format(self):
        self.assertTrue(validate_month_format("2025-01"))
        self.assertTrue(validate_month_format("2025-12"))
        self.assertFalse(validate_month_format("2025-13"))
        self.assertFalse(validate_month_format("invalid"))


class TestForecastEngine(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.tmpdir, "test_forecast.xlsx")
        self._create_test_excel()

    def _create_test_excel(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Test"

        months = ["2025-01", "2025-02", "2025-03"]
        headers = ["SOW Name", "SOW Type"] + months
        ws.append(headers)
        ws.append(["Salary", "income", 5000, 5100, 5200])
        ws.append(["Index Fund", "investment", 20000, 21000, 22000])
        wb.save(self.excel_path)

    def test_forecast_sow(self):
        sow_list = load_excel(self.excel_path)
        result = forecast_sow(sow_list[0], forecast_months=3)
        self.assertEqual(len(result), 3)
        self.assertIn("2025-04", result)
        self.assertIn("2025-06", result)
        self.assertGreater(result["2025-04"], 5200)

    def test_forecast_sow_with_override(self):
        sow_list = load_excel(self.excel_path)
        result = forecast_sow(
            sow_list[0], forecast_months=3, growth_rate=0.10
        )
        self.assertGreater(result["2025-04"], 5200)

    def test_forecast_portfolio(self):
        sow_list = load_excel(self.excel_path)
        result = forecast_portfolio(sow_list, forecast_months=3)
        self.assertEqual(len(result), 2)
        self.assertIn("Salary", result)
        self.assertIn("Index Fund", result)

    def test_get_forecast_months(self):
        sow_list = load_excel(self.excel_path)
        months = get_forecast_months(sow_list, forecast_months=3)
        self.assertEqual(len(months), 3)
        self.assertEqual(months[0], "2025-04")
        self.assertEqual(months[-1], "2025-06")


class TestBreakdown(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.tmpdir, "test_breakdown.xlsx")
        self._create_test_excel()

    def _create_test_excel(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Test"

        months = ["2025-01", "2025-02", "2025-03"]
        headers = ["SOW Name", "SOW Type"] + months
        ws.append(headers)
        ws.append(["Salary", "income", 5000, 5100, 5200])
        ws.append(["Index Fund", "investment", 20000, 21000, 22000])
        ws.append(["Credit Card", "credit", -3000, -2800, -2600])
        wb.save(self.excel_path)

    def test_monthly_totals(self):
        sow_list = load_excel(self.excel_path)
        totals = compute_monthly_totals(sow_list)
        self.assertIn("2025-01", totals)
        self.assertEqual(totals["2025-01"], 5000 + 20000 + (-3000))

    def test_percentage_breakdown(self):
        sow_list = load_excel(self.excel_path)
        target_months = ["2025-01", "2025-02"]
        pct = compute_percentage_breakdown(sow_list, target_months)
        self.assertIn("Salary", pct)
        self.assertIn("Index Fund", pct)
        self.assertIn("2025-01", pct["Salary"])

        total_pct_jan = sum(
            pct[sow_name].get("2025-01", 0)
            for sow_name in pct
        )
        self.assertAlmostEqual(total_pct_jan, 100.0, delta=0.01)

    def test_type_breakdown(self):
        sow_list = load_excel(self.excel_path)
        target_months = ["2025-01", "2025-02"]
        pct = compute_type_breakdown(sow_list, target_months)
        self.assertIn("income", pct)
        self.assertIn("investment", pct)
        self.assertIn("credit", pct)

        total_pct_jan = sum(
            pct[stype].get("2025-01", 0)
            for stype in pct
        )
        self.assertAlmostEqual(total_pct_jan, 100.0, delta=0.01)

    def test_asset_vs_liability(self):
        sow_list = load_excel(self.excel_path)
        target_months = ["2025-01", "2025-02"]
        result = compute_asset_vs_liability(sow_list, target_months)
        self.assertIn("assets", result)
        self.assertIn("liabilities", result)
        self.assertIn("net_worth", result)

        self.assertGreater(result["assets"]["2025-01"], 0)
        self.assertGreater(result["liabilities"]["2025-01"], 0)
        net = result["net_worth"]["2025-01"]
        expected = result["assets"]["2025-01"] - result["liabilities"]["2025-01"]
        self.assertEqual(net, expected)

    def test_net_worth_growth(self):
        sow_list = load_excel(self.excel_path)
        target_months = ["2025-01", "2025-02", "2025-03"]
        result = compute_net_worth_growth(sow_list, target_months)
        self.assertIn("total_growth_pct", result)
        self.assertIn("monthly_cagr_pct", result)


class TestPipeline(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.tmpdir, "test_pipeline.xlsx")
        self._create_test_excel()

    def _create_test_excel(self):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Test"

        months = ["2025-01", "2025-02", "2025-03"]
        headers = ["SOW Name", "SOW Type"] + months
        ws.append(headers)
        ws.append(["Salary", "income", 5000, 5100, 5200])
        ws.append(["Index Fund", "investment", 20000, 21000, 22000])
        ws.append(["Credit Card", "credit", -3000, -2800, -2600])
        wb.save(self.excel_path)

    def test_full_pipeline(self):
        result = run_pipeline(self.excel_path, forecast_months=3)
        self.assertIn("sow_summaries", result)
        self.assertIn("type_summary", result)
        self.assertIn("forecast_months", result)
        self.assertIn("forecasts", result)
        self.assertIn("assumptions", result)
        self.assertIn("net_worth_growth", result)

        self.assertEqual(len(result["sow_summaries"]), 3)

        self.assertIn("income", result["type_summary"])
        self.assertIn("investment", result["type_summary"])
        self.assertIn("credit", result["type_summary"])

    def test_pipeline_with_overrides(self):
        result = run_pipeline(
            self.excel_path,
            forecast_months=3,
            growth_overrides={"investment": 0.10},
            contribution_overrides={"investment": 100},
        )
        self.assertIn("investment", result["assumptions"])
        self.assertEqual(
            result["assumptions"]["investment"]["growth_rate"], 0.10
        )
        self.assertEqual(
            result["assumptions"]["investment"]["monthly_contribution"], 100
        )
        self.assertEqual(
            result["assumptions"]["investment"]["source"], "override"
        )


if __name__ == "__main__":
    unittest.main()